import sys
import serial
from io import StringIO
import csv
from openpyxl import Workbook, load_workbook
import datetime
from dateutil.parser import parse
import time

from PyQt5.QtCore import QCoreApplication, Qt, QUrl
from PyQt5.QtGui import QGuiApplication, QIcon
from PyQt5.QtQml import QQmlApplicationEngine
from PyQt5.QtCore import QObject, pyqtSignal, pyqtSlot, QThread
from PyQt5.Qt import QStringListModel





ampHourvalue=0
batteryCapacity=100

#define and open the serial port
# for windows COM port of pi-top TTYS0
#ser=serial.Serial('COM3')
ser=serial.Serial('/dev/ttyS0', 57600)
# Create an instance of the application
app = QGuiApplication(sys.argv)
# Create QML engine
engine = QQmlApplicationEngine()



class Dash(QObject):

    #create the signals that change the GUI
    auxVoltage = pyqtSignal(float, arguments=['auxvolt'])
    mainVoltage = pyqtSignal(float, arguments=['mainvolt'])
    mainCurrent = pyqtSignal(float, arguments=['maincur'])
    carSpeed = pyqtSignal(float, arguments=['speedmph'])
    ampHour = pyqtSignal(float, arguments=['amphr'])

    @pyqtSlot()
    def reset(self):
        global ampHourvalue
        global batteryCapacity
        ampHourvalue=batteryCapacity #put amp hour capacity here
        self.ampHour.emit(ampHourvalue)


    @pyqtSlot(str)
    def set(self,value):   #allows the user to update the amphour counter
        global ampHourvalue  #pulls in global variable
        ampHourvalue=float(value)  #converts value from string to number
        self.ampHour.emit(ampHourvalue)  #sends the signal to update display.

# create a dashboard object
dashboard = Dash()




class ThreadClass(QThread):
    # Create the signal
    auxVoltage = pyqtSignal(float, arguments=['auxvolt'])
    mainVoltage = pyqtSignal(float, arguments=['mainvolt'])
    mainCurrent = pyqtSignal(float, arguments=['maincur'])
    carSpeed = pyqtSignal(float, arguments=['speedmph'])
    ampHour = pyqtSignal(float, arguments=['amphr'])




    def __init__(self, parent=None):
        super(ThreadClass, self).__init__(parent)

        #connects the signals from the thread to the object
        self.auxVoltage.connect(dashboard.auxVoltage)
        self.mainVoltage.connect(dashboard.mainVoltage)
        self.mainCurrent.connect(dashboard.mainCurrent)
        self.carSpeed.connect(dashboard.carSpeed)
        self.ampHour.connect(dashboard.ampHour)




    def run(self):
        #This implementation of amp hours needs work
        global ampHourvalue
        global batteryCapacity

        try:
            wb=load_workbook("History.xlsx") #attemps to open the history excel file
        except:
            wb=Workbook() #creates and empty excel workbook if histortory is not found


        WorkSheetName=datetime.date.today() #get todays date

        try:
            ws=wb["%s" %WorkSheetName]
        except:
            ws = wb.create_sheet("%s" %WorkSheetName) #create worksheet with the date as title

        firstrun=True  # logical flag to prevent errors on first run


        ws['A1']="Time"
        ws['B1']="Aux Voltage"
        ws['C1']="Main Voltage"
        ws['D1']="Current"
        ws['E1']="Lat"
        ws['F1']="Lon"
        ws['G1']='Speed'
        ws['F1']="Amphours"





        while True:
            data=ser.readline(120) #read the stream
            data=data.decode() #convert stream from bytes to characters
            data=StringIO(data)#convert a stream of characters into string
            dataset=csv.reader(data, delimiter= ',') #read the CSV string into individual array
            dataset=list(dataset) #convert the array to list
            print(dataset)

            #write to excel log


            #extract individual data points
            timestamp=(dataset[0][0])
            auxvalue=float(dataset[0][1])
            mainvalue=float(dataset[0][2])
            currentvalue=float(dataset[0][3])
            speedvalue=float(dataset[0][6])




            auxvalue=round(auxvalue,2)
            mainvalue=round(mainvalue,2)
            currentvalue=round(currentvalue,2)
            speedvalue=round(speedvalue,2)


            # Emit the signals
            self.auxVoltage.emit(auxvalue)
            self.mainVoltage.emit(mainvalue)
            self.mainCurrent.emit(currentvalue)
            self.carSpeed.emit(speedvalue)


            if firstrun:
                firstrun=False
                timestamp=parse(timestamp)

            #calc time between last two records
            else:
                timestamp=parse(timestamp)
                timedelta=timestamp-oldtimestamp
                #calculate the number of amphours used on this interval
                currentused= (currentvalue)*(float(timedelta.seconds)/3600)
                #subtract the number of amphours used from the remaining amount
                ampHourvalue = ampHourvalue - currentused
                if ampHourvalue > batteryCapacity:
                    ampHourvalue = batteryCapacity

                ampHourvalue=round(ampHourvalue,2)
                self.ampHour.emit(ampHourvalue)


            oldtimestamp=timestamp  #saves the current time stamp for use in the next calculation

            dataset[0].append(ampHourvalue)   #adds the calculated amphout value to the dataset
            ws.append(dataset[0])     #appends the current times dataset to the spreedsheet
            wb.save('History.xlsx')    #saves the spreadsheet


        pass








    # register it in the context of QML
    engine.rootContext().setContextProperty("dashboard", dashboard)
    #load qml file into engine
    engine.load(QUrl('main2.qml'))



threadclass=ThreadClass()
threadclass.start()

engine.quit.connect(app.quit)
sys.exit(app.exec_())
app.exec_()
